"""
헤비듀티 FW 콘텐츠 캘린더 xlsx 생성
- Sheet 1: 카테고리 안내 (메인·서브 추가 이유)
- Sheet 2: 발행 일정표 (대표님 양식 + 6콘텐츠)
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date, timedelta

OUT = r'C:\Users\wwkfl\Desktop\heavyduty-content-site\reports\content-calendar\콘텐츠_캘린더.xlsx'

# ============================================
# 카테고리 데이터
# ============================================
CATEGORIES = [
    # 시즌 메인 11개 — 특정 발행 주차 (청바지·모자는 별도 포시즌 박스로)
    (1,  '플란넬·체크셔츠',      date(2026, 8, 17),  '9/3',    'C 서브',     '정점 -2주', date(2026, 9, 3)),
    (2,  '마운틴파카 (60/40)',   date(2026, 9, 14),  '9/28',   'A 시즌메인', '대표님 안 · 60/40 환절기 정점 -2주', date(2026, 9, 28)),
    (3,  '바람막이+아노락',      date(2026, 9, 21),  '10/1',   'A 시즌메인', '정점 -1주 (마운틴파카와 어긋남)', date(2026, 10, 1)),
    (4,  '청자켓·데님자켓',      date(2026, 9, 28),  '10/7',   'A 시즌메인', '정점 -1주 (바람막이와 어긋남)', date(2026, 10, 7)),
    (5,  '필드·헌팅자켓',        date(2026, 10, 5),  '환절기',  'C 메인추가', '🥈 가을 환절기 자유 배치', None),
    (6,  '맨투맨·후드티',        date(2026, 10, 12), '10/26',  'C 메인추가', '🥇 정점 -2주', date(2026, 10, 26)),
    (7,  '플리스',               date(2026, 10, 26), '11/9',   'A 시즌메인', '정점 -2주', date(2026, 11, 9)),
    (8,  '패딩조끼',             date(2026, 11, 2),  '11/16',  'A 시즌메인', '정점 -2주', date(2026, 11, 16)),
    (9,  '코위찬·헤비스웨터',    date(2026, 11, 9),  '11/28',  'C 서브',     '정점 -3주 (어긋남)', date(2026, 11, 28)),
    (10, '패딩',                 date(2026, 11, 16), '12/1',   'A 시즌메인', '정점 -2주', date(2026, 12, 1)),
    (11, '코듀로이',             date(2026, 11, 23), '12/1',   'C 서브',     '정점 -1주 (패딩과 어긋남)', date(2026, 12, 1)),
]

# 포시즌 카테고리 — 발행 주차 없이 매주 자유 발행
POSTSEASON_CATEGORIES = [
    ('청바지 (데님)', '데님 결로 1년 내내 운영. 청자켓·청바지 시너지. 가을 서브 정점 10/1·봄 메인 3월에 추가 푸시 가능.'),
    ('모자',          '빈티지 캡 라인업 (영석대표님 28종 후보). 사입 입고 시점에 맞춰 자유 운영.'),
]

# v13 주차 함수 — 1-7=1주차, 8-14=2주차, 15-21=3주차, 22~말일=4주차
def date_to_week_label(d):
    if d is None:
        return '연중 (자유)'
    day = d.day
    if day <= 7:
        w = 1
    elif day <= 14:
        w = 2
    elif day <= 21:
        w = 3
    else:
        w = 4
    return f'{d.month}월 {w}주차'

# 주차 + 그 주의 월요일~일요일 날짜 범위
def date_to_week_label_with_range(d):
    from datetime import timedelta
    if d is None:
        return '연중 (자유)'
    weekday = d.weekday()  # 0=월, 6=일
    monday = d - timedelta(days=weekday)
    sunday = monday + timedelta(days=6)
    day = d.day
    if day <= 7:
        w = 1
    elif day <= 14:
        w = 2
    elif day <= 21:
        w = 3
    else:
        w = 4
    return f'{d.month}월 {w}주차\n({monday.month}/{monday.day}~{sunday.month}/{sunday.day})'

# 단계별 시점 계산 (정점 기준)
def stage_dates(peak):
    """정점 일자에서 각 단계 시점 계산"""
    if peak is None:
        return None
    from datetime import timedelta
    return {
        'plan':    peak - timedelta(weeks=12),
        'sourcing': peak - timedelta(weeks=8),
        'care':    peak - timedelta(weeks=5),
        'creative': peak - timedelta(weeks=3),
        'launch':  peak - timedelta(weeks=2),
        'peak':    peak,
        'remind':  peak + timedelta(weeks=1),
    }

# 6콘텐츠 패턴 (요일·콘텐츠 유형·일자 오프셋)
CONTENT_PATTERN = [
    (0, '월', '제품 업데이트'),
    (1, '화', '제품 콘텐츠'),
    (2, '수', '정보성 콘텐츠'),
    (3, '목', '문화 콘텐츠'),
    (4, '금', 'OOTD 카드뉴스'),
    (6, '일', 'OOTD 릴스'),
]

# 카테고리별 ref / 메모
REFERENCES = {
    '청자켓·데님자켓': {
        '제품 콘텐츠': 'ref: instagram.com/p/DW6GF_vlIt4',
        '정보성 콘텐츠': 'ref: instagram.com/p/DVJV25CARHD',
        '문화 콘텐츠': '데님자켓 명배우 — 말론 브란도, 제임스 딘, 스티브 맥퀸',
        'OOTD 카드뉴스': 'ref: instagram.com/p/DYlTNnElDqG',
        'OOTD 릴스': 'ref1: instagram.com/p/DYZrizNRHQj / ref2: instagram.com/p/DXWtY1XAWot',
    },
    '바람막이+아노락': {
        '제품 콘텐츠': 'ref: instagram.com/p/DVJV25CARHD',
        '정보성 콘텐츠': 'ref: instagram.com/p/DYJwipXEatd → "빈티지로 사야 하는 바람막이+아노락 브랜드" 변형',
        '문화 콘텐츠': '아노락 흐름 — 이누이트 원조 → 1911년 아문센 남극 정복 → 70s 등산 → 90s 오아시스·갤러거 형제 브릿팝',
        'OOTD 카드뉴스': 'ref: instagram.com/p/DYlTNnElDqG',
        'OOTD 릴스': 'ref1: instagram.com/p/DYZrizNRHQj / ref2: instagram.com/p/DXWtY1XAWot',
    },
    '플리스': {
        '정보성 콘텐츠': '"플리스 원조는 파타고니아가 아니다" — 헬리한센 1961',
        '문화 콘텐츠': '노르웨이 벌목꾼 1961 · 파타고니아 80s 캘리포니아 클라이밍 문화',
    },
    '패딩조끼': {
        '정보성 콘텐츠': '패딩베스트 헤리티지 — 에디바우어 시작',
        '문화 콘텐츠': '90s 힙합 올드머니 룩 · 80s 등산 붐',
    },
    '마운틴파카': {
        '제품 콘텐츠': 'ref: instagram.com/p/DVJV25CARHD',
        '정보성 콘텐츠': 'ref: instagram.com/p/DYJwipXEatd → 마운틴파카 60/40 변형',
        '문화 콘텐츠': '70s 미국 백패킹 붐 · 시에라 클럽 · 60/40 코튼·나일론 헤리티지',
        'OOTD 카드뉴스': 'ref: instagram.com/p/DYlTNnElDqG',
        'OOTD 릴스': 'ref1: instagram.com/p/DYZrizNRHQj / ref2: instagram.com/p/DXWtY1XAWot',
    },
    '패딩': {
        '정보성 콘텐츠': '다운파카 발달사 — 에디바우어 Skyliner (1936)',
        '문화 콘텐츠': '80s 노르웨이 등산가 · 에베레스트 탐험대 · 80s 다운 붐',
    },
    '모자 (포시즌)': {
        '정보성 콘텐츠': '캡·트럭커캡·6패널 헤리티지',
        '문화 콘텐츠': '영화·뮤지션 데일리 모자 — 폴 뉴먼, 잭 케루악, 70s 트럭 드라이버',
    },
    '청바지 (데님, 포시즌)': {
        '정보성 콘텐츠': '리바이스 505 / 559 / 569 모델별 차이',
        '문화 콘텐츠': '리바이스 505 록밴드 — 롤링 스톤즈, 라몬즈, 도어즈, 비치 보이즈',
    },
    '맨투맨·후드티': {
        '정보성 콘텐츠': '챔피언 · 러셀 · 헤인즈 빈티지 헤리티지',
        '문화 콘텐츠': '록키 (1976) · 풋볼 컬리지 시대 · 90s 힙합 (NBA 빅 사이즈)',
    },
    '필드·헌팅자켓': {
        '정보성 콘텐츠': 'M-65 · OG-107 야상 · 영국군 P64 — 군복 헤리티지',
        '문화 콘텐츠': '모드족 · 펑크 신 · 영화 (택시 드라이버, 풀 메탈 자켓)',
    },
    '코위찬·헤비스웨터': {
        '정보성 콘텐츠': '코위찬 부족(Cowichan) 손뜨개 · 피셔맨 스웨터 헤리티지',
        '문화 콘텐츠': '캐나다 원주민 부족 문화 · 영화 "빅 르보스키" 더 듀드 가디건',
    },
    '플란넬·체크셔츠': {
        '정보성 콘텐츠': '플란넬 헤리티지 — 필슨·펜들턴·울리치·LL빈 (미국 헤리티지 워크웨어)',
        '문화 콘텐츠': '90s 그런지 신 — 너바나·펄잼·사운드가든 / 영화 — 파이트 클럽 (1999), 캡틴 판타스틱 (2016) / 미국 농부·트럭 운전사 데일리',
    },
    '코듀로이': {
        '정보성 콘텐츠': '코듀로이 소재 발달사 — 영국 워크웨어에서 미국 60s~70s 캐주얼로',
        '문화 콘텐츠': '영화 — 더 로열 테넌바움, 라이프 어쿼틱 (웨스 앤더슨), 보이후드 (2014), 라스트 픽처 쇼 (1971) / 매튜 매커너히 70s 룩',
    },
}

# C 카테고리 추가 이유
C_REASONS = {
    'main_intro': '대표님 원본 6개 카테고리(A) + 포시즌 2개(B) 외에 추가된 C 카테고리들의 추가 이유',
    'rank_1': {
        'name': '🥇 1순위 — 맨투맨·후드티 (메인 추가)',
        'why': [
            '검색량 최고 (정점 100+, 청자켓의 약 2배)',
            '매출 잠재력 최강 — 회전 빠름, 가격대 합리적, 일상 옷',
            '챔피언·러셀·헤인즈 빈티지 = 미국 스포츠웨어 헤리티지',
            '10월 중하순 빈 기간 자연스럽게 채움 (가을 정점 직후)',
            '콘텐츠 풍부 — 70s~90s 록·운동·영화 시대성',
        ]
    },
    'rank_2': {
        'name': '🥈 2순위 — 필드·헌팅자켓 (메인 추가)',
        'why': [
            '헤비듀티 정체성 가장 강함 — M-65, 미군 OG-107, 영국군 P64',
            '기획서 "쉘 챕터(벤타일·60/40·고어텍스)"와 직접 연결',
            'A의 청자켓·바람막이 시즌 보완 — 가을 환절기 라이트 아우터 결',
            '콘텐츠 무한 — 밀리터리 스토리, 모드족, 펑크 신, 영화',
            '대표님 본인이 예시로 언급한 카테고리',
        ]
    },
    'sub_1': {
        'name': '코위찬·헤비스웨터 (서브)',
        'why': [
            '패딩 시즌(12월)과 시너지 — 겨울 헤비 보온 결',
            '정점 11/4주차로 패딩(12/1) 직전 시점에 보완',
            '코위찬 부족 문화 · 빅 르보스키 더 듀드 — 콘텐츠 차별화',
            '객단가 높은 마니아 시장 — 매출 보조',
        ]
    },
    'sub_2': {
        'name': '코듀로이 (서브)',
        'why': [
            '패딩과 같은 12월 시즌 — 겨울 헤비웨어 결',
            '자켓 + 팬츠 두 가지로 매출 다각화',
            '시즌 마무리 시점에 자연스러운 클로징 카테고리',
            '70s 영화·웨스 앤더슨 등 콘텐츠 풍부',
        ]
    }
}

# ============================================
# 스타일
# ============================================
def thin_border():
    return Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

def thick_border():
    return Border(left=Side(style='medium'), right=Side(style='medium'),
                  top=Side(style='medium'), bottom=Side(style='medium'))

# ============================================
# Workbook 작성
# ============================================
wb = Workbook()

# ========== Sheet 1: 카테고리 안내 ==========
ws1 = wb.active
ws1.title = '카테고리 안내'

ws1['A1'] = '헤비듀티 아카이브 · FW 콘텐츠 캘린더'
ws1['A1'].font = Font(name='맑은 고딕', size=20, bold=True)
ws1.merge_cells('A1:F1')

ws1['A2'] = '카테고리 구성 안내 — 시즌 카테고리 11개 (정점 -2주 발행) + 포시즌 2개 (청바지·모자, 매주 자유)'
ws1['A2'].font = Font(name='맑은 고딕', size=11, color='525252')
ws1.merge_cells('A2:F2')

row = 4
# 헤더
ws1.cell(row=row, column=1, value='구분').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ws1.cell(row=row, column=2, value='카테고리').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ws1.cell(row=row, column=3, value='발행 주차').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ws1.cell(row=row, column=4, value='정점').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ws1.cell(row=row, column=5, value='역할').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ws1.cell(row=row, column=6, value='비고').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
for col in range(1, 7):
    c = ws1.cell(row=row, column=col)
    c.fill = PatternFill('solid', fgColor='000000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border()
row += 1

for cat in CATEGORIES:
    seq, name, week_date, peak, role, note, peak_date = cat
    ws1.cell(row=row, column=1, value=str(seq))
    ws1.cell(row=row, column=2, value=name)
    ws1.cell(row=row, column=3, value=date_to_week_label(week_date))
    ws1.cell(row=row, column=4, value=peak)
    ws1.cell(row=row, column=5, value=role)
    ws1.cell(row=row, column=6, value=note)
    # 역할에 따라 색 다르게 — A 시즌메인 + C 메인추가 = 진한 주황 (확정 강조)
    if 'A' in role:
        fill_color = 'FFB84D'  # 진한 주황
        is_bold = True
    elif 'C 메인' in role:
        fill_color = 'FFB84D'  # 진한 주황
        is_bold = True
    elif 'B' in role:
        fill_color = 'E8F0FF'
        is_bold = False
    else:
        fill_color = 'F5F5F5'  # C 서브 = 연회색
        is_bold = False
    for col in range(1, 7):
        c = ws1.cell(row=row, column=col)
        c.font = Font(name='맑은 고딕', size=10, bold=is_bold)
        c.fill = PatternFill('solid', fgColor=fill_color)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = thin_border()
    row += 1

row += 2

# 포시즌 카테고리 박스 (청바지·모자)
ws1.cell(row=row, column=1, value='B. 포시즌 카테고리 — 발행 주차 외 매주 자유 운영')
ws1.cell(row=row, column=1).font = Font(name='맑은 고딕', size=13, bold=True, color='FFFFFF')
ws1.cell(row=row, column=1).fill = PatternFill('solid', fgColor='4A6FA5')  # 차분한 파란색
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws1.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws1.row_dimensions[row].height = 26
row += 1

# 포시즌 박스 — 추신 메시지
ws1.cell(row=row, column=1, value='⤷ 위 시즌 카테고리 외에 청바지·모자는 매주 적절한 수량을 콘텐츠에 자연스럽게 포함시킬 예정입니다.')
ws1.cell(row=row, column=1).font = Font(name='맑은 고딕', size=11, bold=True, color='000000')
ws1.cell(row=row, column=1).fill = PatternFill('solid', fgColor='E8F0FF')
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws1.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws1.row_dimensions[row].height = 24
row += 1

# 포시즌 헤더
ws1.cell(row=row, column=1, value='카테고리').font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
ws1.cell(row=row, column=2, value='운영 방식').font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
ws1.cell(row=row, column=3, value='메모').font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
for col in range(1, 7):
    c = ws1.cell(row=row, column=col)
    c.fill = PatternFill('solid', fgColor='000000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border()
ws1.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
row += 1

for cat_name, note in POSTSEASON_CATEGORIES:
    ws1.cell(row=row, column=1, value=cat_name)
    ws1.cell(row=row, column=2, value='매주 적절한 수량 포함')
    ws1.cell(row=row, column=3, value=note)
    ws1.merge_cells(start_row=row, start_column=3, end_row=row, end_column=6)
    for col in range(1, 7):
        c = ws1.cell(row=row, column=col)
        c.font = Font(name='맑은 고딕', size=10)
        c.fill = PatternFill('solid', fgColor='E8F0FF')
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
        c.border = thin_border()
    ws1.row_dimensions[row].height = 30
    row += 1

row += 2

# C 카테고리 추가 이유
ws1.cell(row=row, column=1, value='C 카테고리 추가 이유')
ws1.cell(row=row, column=1).font = Font(name='맑은 고딕', size=14, bold=True)
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
row += 1

ws1.cell(row=row, column=1, value=C_REASONS['main_intro'])
ws1.cell(row=row, column=1).font = Font(name='맑은 고딕', size=10, color='525252')
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
row += 2

for key in ['rank_1', 'rank_2', 'sub_1', 'sub_2']:
    item = C_REASONS[key]
    # 카테고리 제목
    ws1.cell(row=row, column=1, value=item['name'])
    ws1.cell(row=row, column=1).font = Font(name='맑은 고딕', size=12, bold=True, color='FFFFFF')
    ws1.cell(row=row, column=1).fill = PatternFill('solid', fgColor='000000')
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws1.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
    row += 1
    # 이유 리스트
    for reason in item['why']:
        ws1.cell(row=row, column=1, value='  • ' + reason)
        ws1.cell(row=row, column=1).font = Font(name='맑은 고딕', size=10)
        ws1.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 1
    row += 1

# ============================================
# C 카테고리 검토 후보 풀 — 담당자 의견 (별도 섹션)
# ============================================
row += 3
ws1.cell(row=row, column=1, value='C 카테고리 검토 풀 — 담당자 추천')
ws1.cell(row=row, column=1).font = Font(name='맑은 고딕', size=15, bold=True)
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
row += 1

ws1.cell(row=row, column=1, value='메인 12개 외에 검토한 카테고리 후보들. 시기·결·콘텐츠·매출·매장 인사이트 종합 검토 후 채택 결정.')
ws1.cell(row=row, column=1).font = Font(name='맑은 고딕', size=10, color='525252')
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
row += 2

# 후보 풀 표 헤더
candidate_headers = ['후보', '발행 시기 / 정점', '헤비듀티 결', '콘텐츠 잠재력', '매출·회전', '채택']
for col, h in enumerate(candidate_headers, 1):
    c = ws1.cell(row=row, column=col, value=h)
    c.font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='000000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border()
row += 1

# 검토했던 모든 후보 (채택 + 미채택)
candidates = [
    ('플란넬·체크셔츠 ⭐', '가을 환절기 / 9/3 ✓ 검증', '★★★★ 헤리티지 워크웨어 (필슨·펜들턴·울리치)', '★★★★★ 90s 그런지·영화·미국 농부', '★★★★ 데일리 회전 빠름', '✓ 채택 (①)'),
    ('맨투맨·후드티 ⭐', '가을 메인 / 10/26 ✓ 검증', '★★★★ 챔피언·러셀 빈티지', '★★★★★ 록키·90s 힙합·풋볼', '★★★★ 검색량 최고', '✓ 채택 (⑨)'),
    ('필드·헌팅자켓 ⭐', '봄 메인 / 환절기 ✓ 검증', '★★★★★ M-65·미군·기획서 쉘 챕터 연결', '★★★★ 영화 다수 (택시 드라이버·풀 메탈)', '★★★ 객단가 큼', '✓ 채택 (⑩)'),
    ('코위찬·헤비스웨터 ⭐', '겨울 / 11/28 ✓ 검증', '★★★ 캐나다 원주민 헤리티지', '★★★ 빅 르보스키·노르웨이 어부', '★★★ 객단가 큼', '✓ 채택 (⑪)'),
    ('코듀로이 ⭐', '늦가을~겨울 / 12/1 ✓ 검증', '★★★ 70s 워크웨어', '★★★ 70s 영화·웨스 앤더슨', '★★★ 자켓+팬츠 다각화', '✓ 채택 (⑫)'),
    ('레터맨 자켓 (바시티)', '가을~초겨울 / 데이터 없음', '★★★ 50s~70s 미국 컬리지', '★★★★ 영화 그리스·아메리칸 그래피티', '★★★ 객단가 큼', '✗ 미채택'),
    ('워크팬츠·치노·카고', '봄·환절기 / 2/4 (신뢰도 낮음)', '★★★★ 헤리티지 워크 (디키즈·벤데이비스)', '★★★ 일상 옷·90s 스케이트', '★★★ 회전 빠름', '✗ 시기 불일치'),
    ('CPO 셔츠 (헤비 울 체크)', '늦가을~겨울 / 데이터 없음', '★★★ 해군 헤리티지', '★★★ 50s 미국 해군·그런지', '★★★ 마니아 시장', '✗ 풀 좁음'),
    ('데크자켓·N-1 / N-3B', '겨울 / 데이터 없음', '★★★★★ 미해군 밀리터리 헤리티지', '★★★ 영화 진주만·탑건', '★★★ 객단가 매우 큼', '✗ 풀 좁음·필드자켓과 결 중복'),
    ('메리노 울 헨리넥·터틀넥', '가을~겨울 베이스 / 데이터 없음', '★★★ 19세기 노동자 헤리티지', '★★ 잭 케루악·비트 제너레이션', '★★ 객단가 낮음', '✗ 풀 좁음'),
    ('레더자켓·라이딩자켓', '겨울 / 데이터 없음', '★★ 모터사이클·바이커 (헤비듀티 결 다름)', '★★★★ 와일드 원·람보·매트릭스', '★★★★ 객단가 매우 큼', '✗ 정체성 결 불일치'),
    ('부츠·헤리티지 슈즈', '겨울 / 데이터 없음', '★★★★ 레드윙·닥마틴·팀버랜드', '★★★ 작업화 헤리티지', '★★★★ 객단가 매우 큼', '✗ 액세서리 (별도 트랙)'),
]

for cand in candidates:
    is_recommended = '⭐' in cand[0]
    is_adopted = '✓ 채택' in cand[5]
    if is_adopted:
        fill_color = 'FFF4E0'  # 채택 노란색
    else:
        fill_color = 'F5F5F5'  # 미채택 회색
    for col, val in enumerate(cand, 1):
        c = ws1.cell(row=row, column=col, value=val)
        c.font = Font(name='맑은 고딕', size=10, bold=is_adopted)
        c.fill = PatternFill('solid', fgColor=fill_color)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws1.row_dimensions[row].height = 32
    row += 1

# 담당자 추천 박스 — 어필
row += 2
ws1.cell(row=row, column=1, value='🎯  담당자 의견 — "저는 이런 카테고리를 추천합니다"')
ws1.cell(row=row, column=1).font = Font(name='맑은 고딕', size=14, bold=True, color='FFFFFF')
ws1.cell(row=row, column=1).fill = PatternFill('solid', fgColor='000000')
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws1.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws1.row_dimensions[row].height = 28
row += 1

ws1.cell(row=row, column=1, value='담당자 최종 선택 — 플란넬·체크셔츠 (시즌 시작 ① 자리)')
ws1.cell(row=row, column=1).font = Font(name='맑은 고딕', size=12, bold=True, color='000000')
ws1.cell(row=row, column=1).fill = PatternFill('solid', fgColor='FFF4E0')
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws1.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws1.row_dimensions[row].height = 24
row += 1

recommendation_reasons = [
    '【 매장 운영 인사이트 】 기존에 사입했던 체크셔츠는 수량은 적었지만 빠르게 판매되었던 이력이 있습니다. 풀을 더 키워서 운영해도 충분히 회전될 카테고리로 판단됩니다.',
    '【 데이터 】 우리 v13 분석에서 검증된 카테고리. 9/3 가을 정점 기준 발행 일정 근거 명확.',
    '【 정체성 】 미국 헤리티지 워크웨어 (필슨·펜들턴·울리치·LL빈) — 헤비듀티 결과 정확히 부합.',
    '【 콘텐츠 】 90s 그런지(너바나·펄잼·사운드가든), 미국 농부·트럭 운전사, 영화(파이트 클럽·캡틴 판타스틱) — 콘텐츠 풀 매우 풍부.',
    '【 매출 】 객단가 적정 + 데일리 회전 빠름. 패딩(객단가 큼·회전 느림) 보완 효과.',
]

for reason in recommendation_reasons:
    ws1.cell(row=row, column=1, value=reason)
    ws1.cell(row=row, column=1).font = Font(name='맑은 고딕', size=10.5)
    ws1.cell(row=row, column=1).fill = PatternFill('solid', fgColor='FFF4E0')
    ws1.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws1.row_dimensions[row].height = 28
    row += 1

# 컬럼 너비
ws1.column_dimensions['A'].width = 22
ws1.column_dimensions['B'].width = 24
ws1.column_dimensions['C'].width = 22
ws1.column_dimensions['D'].width = 32
ws1.column_dimensions['E'].width = 18
ws1.column_dimensions['F'].width = 16

# ========== Sheet 2: 발행 일정 ==========
ws2 = wb.create_sheet('발행 일정')

ws2['A1'] = '헤비듀티 아카이브 · FW 발행 일정표'
ws2['A1'].font = Font(name='맑은 고딕', size=18, bold=True)
ws2.merge_cells('A1:F1')

ws2['A2'] = '공통 사입 일괄 진행 · 주간 패턴: 월(제품)→화(제품콘텐츠)→수(정보성)→목(문화 NEW)→금(OOTD 카드뉴스)→일(OOTD 릴스)'
ws2['A2'].font = Font(name='맑은 고딕', size=10, color='525252')
ws2.merge_cells('A2:F2')

# 헤더
row = 4
headers = ['완료', '날짜', '요일', '카테고리', '콘텐츠 유형', '레퍼런스 / 메모']
for col, h in enumerate(headers, 1):
    c = ws2.cell(row=row, column=col, value=h)
    c.font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='000000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border()
row += 1

# 카테고리별 행
for cat in CATEGORIES:
    seq, name, week_start, peak, role, _, peak_date = cat
    # 카테고리 헤더 행
    label = f'{["①","②","③","④","⑤","⑥","⑦","⑧","⑨","⑩","⑪"][seq-1]} {name}  ·  {date_to_week_label(week_start)}  ·  정점 {peak}  ·  {role}'
    ws2.cell(row=row, column=1, value=label)
    ws2.cell(row=row, column=1).font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
    ws2.cell(row=row, column=1).fill = PatternFill('solid', fgColor='404040')
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws2.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
    ws2.cell(row=row, column=1).border = thin_border()
    row += 1

    # 6콘텐츠 행
    for offset, day_kor, content_type in CONTENT_PATTERN:
        content_date = week_start + timedelta(days=offset)
        ref = REFERENCES.get(name, {}).get(content_type, '')

        ws2.cell(row=row, column=1, value='☐')
        ws2.cell(row=row, column=2, value=content_date.strftime('%m/%d'))
        ws2.cell(row=row, column=3, value=day_kor)
        ws2.cell(row=row, column=4, value=name)
        ws2.cell(row=row, column=5, value=content_type)
        ws2.cell(row=row, column=6, value=ref)

        for col in range(1, 7):
            c = ws2.cell(row=row, column=col)
            c.font = Font(name='맑은 고딕', size=10)
            c.alignment = Alignment(horizontal='center' if col in [1,2,3] else 'left', vertical='center', wrap_text=True)
            c.border = thin_border()
            # 목요일(문화)은 강조
            if content_type == '문화 콘텐츠':
                c.fill = PatternFill('solid', fgColor='FFF9E0')
        row += 1

# 컬럼 너비
ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 9
ws2.column_dimensions['C'].width = 7
ws2.column_dimensions['D'].width = 22
ws2.column_dimensions['E'].width = 16
ws2.column_dimensions['F'].width = 60

# 행 높이
for r in range(1, row):
    ws2.row_dimensions[r].height = 20

# ========== Sheet 3: 단계별 데드라인 ==========
ws_dl = wb.create_sheet('단계별 데드라인')

ws_dl['A1'] = '단계별 데드라인 — 12개 카테고리 정점 -2주 발행 기준'
ws_dl['A1'].font = Font(name='맑은 고딕', size=16, bold=True)
ws_dl.merge_cells('A1:I1')

ws_dl['A2'] = '각 카테고리의 정점을 기준으로 단계별 데드라인 계산. 데드라인은 마감일 기준 — 그 시점까지 끝내야 함.'
ws_dl['A2'].font = Font(name='맑은 고딕', size=10, color='525252')
ws_dl.merge_cells('A2:I2')

# 헤더
dl_row = 4
dl_headers = ['#', '카테고리', '정점', '① 캘린더 확정\n(P-12주)', '② 사입 마감\n(P-8주)', '③ 제품 케어\n(P-5주)', '④ 콘텐츠 기획·제작\n(P-3주)', '⑤ 발행 (6콘텐츠)\n(P-2주)', '⑥ 리마인드\n(P+1주)']
for col, h in enumerate(dl_headers, 1):
    c = ws_dl.cell(row=dl_row, column=col, value=h)
    c.font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='000000')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_border()
ws_dl.row_dimensions[dl_row].height = 50
dl_row += 1

for cat in CATEGORIES:
    seq, name, week_start, peak, role, _, peak_date = cat
    stages = stage_dates(peak_date)

    ws_dl.cell(row=dl_row, column=1, value=seq)
    ws_dl.cell(row=dl_row, column=2, value=name)
    ws_dl.cell(row=dl_row, column=3, value=peak)

    # 역할에 따라 색 다르게 (확정 = A·C메인 = 주황)
    if 'A' in role or 'C 메인' in role:
        fill_color = 'FFB84D'
        is_confirmed = True
    elif 'B' in role:
        fill_color = 'E8F0FF'
        is_confirmed = False
    else:
        fill_color = 'F5F5F5'
        is_confirmed = False

    if stages:
        ws_dl.cell(row=dl_row, column=4, value=date_to_week_label_with_range(stages['plan']))
        ws_dl.cell(row=dl_row, column=5, value=date_to_week_label_with_range(stages['sourcing']))
        ws_dl.cell(row=dl_row, column=6, value=date_to_week_label_with_range(stages['care']))
        ws_dl.cell(row=dl_row, column=7, value=date_to_week_label_with_range(stages['creative']))
        ws_dl.cell(row=dl_row, column=8, value=date_to_week_label_with_range(stages['launch']))
        ws_dl.cell(row=dl_row, column=9, value=date_to_week_label_with_range(stages['remind']))

        for col in range(1, 10):
            c = ws_dl.cell(row=dl_row, column=col)
            c.font = Font(name='맑은 고딕', size=9, bold=is_confirmed or (col == 8))
            c.fill = PatternFill('solid', fgColor='FFFFCC' if col == 8 else fill_color)  # 발행만 노란 강조
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin_border()
    else:
        # peak_date 없는 카테고리 (필드자켓 같은 환절기) — 셀 병합 + 안내
        for col in range(1, 4):
            c = ws_dl.cell(row=dl_row, column=col)
            c.font = Font(name='맑은 고딕', size=9, bold=is_confirmed)
            c.fill = PatternFill('solid', fgColor=fill_color)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin_border()
        # 단계별 6칸 (D~I) 병합 + 안내 텍스트
        guide_text = '청자켓과 비슷한 시기, 컨셉트 공유하는 제품군 — 10월 3주차 빈 구간 추천. 정점 데이터 없어 자유 배치.'
        ws_dl.cell(row=dl_row, column=4, value=guide_text)
        ws_dl.merge_cells(start_row=dl_row, start_column=4, end_row=dl_row, end_column=9)
        c = ws_dl.cell(row=dl_row, column=4)
        c.font = Font(name='맑은 고딕', size=10, italic=True, color='525252')
        c.fill = PatternFill('solid', fgColor='FFF4E0')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()

    ws_dl.row_dimensions[dl_row].height = 45  # 날짜 범위 추가로 줄 늘어남
    dl_row += 1

# 안내 박스
dl_row += 2
ws_dl.cell(row=dl_row, column=1, value='💡 데드라인 운영 원칙')
ws_dl.cell(row=dl_row, column=1).font = Font(name='맑은 고딕', size=13, bold=True, color='FFFFFF')
ws_dl.cell(row=dl_row, column=1).fill = PatternFill('solid', fgColor='000000')
ws_dl.merge_cells(start_row=dl_row, start_column=1, end_row=dl_row, end_column=9)
ws_dl.cell(row=dl_row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws_dl.row_dimensions[dl_row].height = 26
dl_row += 1

principles = [
    '• 데드라인은 마감일 — 그 시점까지 끝나야 한다는 의미. 더 빠르면 OK.',
    '• 캘린더가 확정되면 사입·제품 케어는 가능한 한 빠르게 진행 → 콘텐츠 기획·제작 시간 확보.',
    '• 발행 (⑤)은 정점 -2주에 시작 → 5콘텐츠(월·화·수·금·일) + 문화 콘텐츠(목) 한 주에 6콘텐츠 다 발행.',
    '• 리마인드(⑥)는 정점 직후 1~2주 안에 OOTD 또는 정보성 1콘텐츠 추가.',
    '• 모자(포시즌)·청바지(포시즌)는 시즌 시퀀스와 무관하게 사입 입고 시점에 맞춰 자유롭게 운영.',
]

for p in principles:
    ws_dl.cell(row=dl_row, column=1, value=p)
    ws_dl.cell(row=dl_row, column=1).font = Font(name='맑은 고딕', size=10.5)
    ws_dl.cell(row=dl_row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    ws_dl.merge_cells(start_row=dl_row, start_column=1, end_row=dl_row, end_column=9)
    ws_dl.row_dimensions[dl_row].height = 22
    dl_row += 1

# 컬럼 너비
ws_dl.column_dimensions['A'].width = 5
ws_dl.column_dimensions['B'].width = 22
ws_dl.column_dimensions['C'].width = 12
for c_letter in ['D', 'E', 'F', 'G', 'H', 'I']:
    ws_dl.column_dimensions[c_letter].width = 17

# ========== Sheet 4: 운영 흐름 예시 (간트차트) ==========
ws_flow = wb.create_sheet('운영 흐름 예시')

ws_flow['A1'] = '운영 흐름 예시 — 청자켓·데님자켓 (정점 10/7 = 10월 1주차)'
ws_flow['A1'].font = Font(name='맑은 고딕', size=16, bold=True)
ws_flow.merge_cells('A1:N1')

ws_flow['A2'] = '정점 대비 각 단계별 시점. 발행 기준(-2주 / -4주 / -5주) 따라 발행 주차가 달라짐.'
ws_flow['A2'].font = Font(name='맑은 고딕', size=10, color='525252')
ws_flow.merge_cells('A2:N2')

# 주차 헤더 (7월 4주차 ~ 10월 4주차)
WEEKS = [
    ('7/4주', '7월\n4주차', 'P-11주'),
    ('8/1주', '8월\n1주차', 'P-10주'),
    ('8/2주', '8월\n2주차', 'P-9주'),
    ('8/3주', '8월\n3주차', 'P-8주'),
    ('8/4주', '8월\n4주차', 'P-7주'),
    ('9/1주', '9월\n1주차', 'P-6주'),
    ('9/2주', '9월\n2주차', 'P-5주'),
    ('9/3주', '9월\n3주차', 'P-4주'),
    ('9/4주', '9월\n4주차', 'P-3주'),
    ('10/1주', '10월\n1주차', 'P-2주'),
    ('10/2주', '10월\n2주차', 'P-1주'),
    ('10/3주', '10월\n3주차', '🎯 정점'),
    ('10/4주', '10월\n4주차', 'P+1주'),
    ('11/1주', '11월\n1주차', 'P+2주'),
]
# 잠깐: 정점이 10/7이면 10월 1주차가 정점 주. P-2주 = 9월 3주차, P-5주 = 8월 4주차. 다시 정리.
WEEKS = [
    ('8/1주', '8월\n1주차', 'P-9주', None),
    ('8/2주', '8월\n2주차', 'P-8주', None),
    ('8/3주', '8월\n3주차', 'P-7주', None),
    ('8/4주', '8월\n4주차', 'P-6주', None),
    ('9/1주', '9월\n1주차', 'P-5주', None),
    ('9/2주', '9월\n2주차', 'P-4주', None),
    ('9/3주', '9월\n3주차', 'P-3주', None),
    ('9/4주', '9월\n4주차', 'P-2주', None),
    ('10/1주', '10월\n1주차', '🎯 정점', None),
    ('10/2주', '10월\n2주차', 'P+1주', None),
    ('10/3주', '10월\n3주차', 'P+2주', None),
    ('10/4주', '10월\n4주차', 'P+3주', None),
]

# 단계별 활동 주차 매핑
# (단계, [활동 주차 인덱스 리스트], 색상, 설명)
STAGES = [
    ('① 콘텐츠 캘린더 확정', [0],          '404040', '카테고리·제품·콘텐츠 앵글 결정 (P-9주 ~ P-12주)'),
    ('② 사입',                [1, 2],       '0070C0', '이베이·거래처 입찰·컨택 (P-8주 ~ P-9주 마감)'),
    ('③ 제품 케어',           [3, 4, 5],    '00B050', '세탁·검수·촬영·매장 등록 (P-5~6주)'),
    ('④ 콘텐츠 기획·제작',    [5, 6, 7],    'FFC000', '서사·카드뉴스·릴스 편집 (P-3~5주)'),
    ('⑤ 발행 옵션 A (-2주)',  [7],          'FF6B6B', '월·화·수·목·금·일 6콘텐츠 (정점 직전 진입)'),
    ('⑤ 발행 옵션 B (-4주)',  [5],          'FFB6B6', '월·화·수·목·금·일 6콘텐츠 (균형)'),
    ('⑤ 발행 옵션 C (-5주)',  [4],          'FFE0E0', '월·화·수·목·금·일 6콘텐츠 (빌드업)'),
    ('🎯 정점 (P)',            [8],          '000000', '검색량 90% 진입 첫날 — 골든타임 시작'),
    ('⑥ 리마인드 콘텐츠',     [9, 10],      'A8A8A8', 'OOTD 또는 정보성 1~2콘텐츠 추가 (정점 직후)'),
]

# 표 헤더 작성
header_row = 4
ws_flow.cell(row=header_row, column=1, value='단계').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ws_flow.cell(row=header_row, column=1).fill = PatternFill('solid', fgColor='000000')
ws_flow.cell(row=header_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws_flow.cell(row=header_row, column=1).border = thin_border()

for i, (date_label, week_label, p_label, _) in enumerate(WEEKS):
    col = i + 2
    ws_flow.cell(row=header_row, column=col, value=f'{week_label}\n({p_label})')
    c = ws_flow.cell(row=header_row, column=col)
    c.font = Font(name='맑은 고딕', size=9, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='000000' if '정점' in p_label else '404040')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_border()

ws_flow.cell(row=header_row, column=len(WEEKS)+2, value='설명').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ws_flow.cell(row=header_row, column=len(WEEKS)+2).fill = PatternFill('solid', fgColor='000000')
ws_flow.cell(row=header_row, column=len(WEEKS)+2).alignment = Alignment(horizontal='center', vertical='center')
ws_flow.cell(row=header_row, column=len(WEEKS)+2).border = thin_border()

ws_flow.row_dimensions[header_row].height = 45

# 단계별 행
for stage_idx, (stage_name, active_weeks, color, desc) in enumerate(STAGES):
    row = header_row + 1 + stage_idx
    # 단계 이름
    c = ws_flow.cell(row=row, column=1, value=stage_name)
    c.font = Font(name='맑은 고딕', size=10, bold=True)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = thin_border()

    # 주차별 셀
    for i in range(len(WEEKS)):
        col = i + 2
        c = ws_flow.cell(row=row, column=col)
        if i in active_weeks:
            c.value = '■'
            c.font = Font(name='맑은 고딕', size=14, bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor=color)
            c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()

    # 설명
    c = ws_flow.cell(row=row, column=len(WEEKS)+2, value=desc)
    c.font = Font(name='맑은 고딕', size=10)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    c.border = thin_border()

    ws_flow.row_dimensions[row].height = 28

# 컬럼 너비
ws_flow.column_dimensions['A'].width = 26
for i in range(len(WEEKS)):
    ws_flow.column_dimensions[get_column_letter(i+2)].width = 10
ws_flow.column_dimensions[get_column_letter(len(WEEKS)+2)].width = 50

# 발행 기준 결정 안내 박스
row = header_row + len(STAGES) + 3
ws_flow.cell(row=row, column=1, value='📌 발행 기준 결정 — 3가지 옵션 비교 (위 표의 ⑤ 발행 옵션 A/B/C 참고)')
ws_flow.cell(row=row, column=1).font = Font(name='맑은 고딕', size=13, bold=True, color='FFFFFF')
ws_flow.cell(row=row, column=1).fill = PatternFill('solid', fgColor='000000')
ws_flow.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(WEEKS)+2)
ws_flow.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws_flow.row_dimensions[row].height = 26
row += 1

# 옵션 비교 표
opt_headers = ['옵션', '발행 시점', '청자켓 발행 주차', '장점', '단점']
for col, h in enumerate(opt_headers, 1):
    c = ws_flow.cell(row=row, column=col, value=h)
    c.font = Font(name='맑은 고딕', size=10, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='000000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border()
ws_flow.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
ws_flow.merge_cells(start_row=row, start_column=9, end_row=row, end_column=len(WEEKS)+2)
row += 1

options = [
    ('A. 정점 -2주', '오픈 시점 (골든타임 시작)', '9월 4주차 (9/21주)', '트래픽 정점 효과 최강. 검색량 정점에 가장 가까이 노출.', '발행 시점이 늦어 시즌 시작감 약함. 사입~제작 압박.'),
    ('B. 정점 -4주', '균형 (빌드업 + 정점 직전)', '9월 2주차 (9/7주)', '시즌 시작 + 정점 둘 다 노출 가능. 안정적.', '특별히 강한 임팩트 없음.'),
    ('C. 정점 -5주', '빌드업 위주', '9월 1주차 (8/31주)', '시즌 초입 노출 강함. 콘텐츠·사입 일정 여유.', '정점에서 멀어 트래픽 정점 직접 효과는 약함. 2차 푸시 필수.'),
]

for opt in options:
    ws_flow.cell(row=row, column=1, value=opt[0])
    ws_flow.cell(row=row, column=2, value=opt[1])
    ws_flow.cell(row=row, column=3, value=opt[2])
    ws_flow.cell(row=row, column=4, value=opt[3])
    ws_flow.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
    ws_flow.cell(row=row, column=9, value=opt[4])
    ws_flow.merge_cells(start_row=row, start_column=9, end_row=row, end_column=len(WEEKS)+2)
    for col in [1, 2, 3, 4, 9]:
        c = ws_flow.cell(row=row, column=col)
        c.font = Font(name='맑은 고딕', size=10)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
        c.border = thin_border()
    ws_flow.row_dimensions[row].height = 32
    row += 1

# ========== Sheet 4: 문화 콘텐츠 가이드 ==========
ws3 = wb.create_sheet('문화 콘텐츠 가이드')

ws3['A1'] = '문화 콘텐츠 — 운영 가이드'
ws3['A1'].font = Font(name='맑은 고딕', size=18, bold=True)
ws3.merge_cells('A1:C1')

ws3['A2'] = '각 제품의 문화적 맥락(인물·영화·음악·시대)을 콘텐츠로 풀어내는 신규 콘텐츠 유형. 매주 목요일 발행.'
ws3['A2'].font = Font(name='맑은 고딕', size=11, color='525252')
ws3.merge_cells('A2:C2')

ws3['A4'] = '콘텐츠 정의'
ws3['A4'].font = Font(name='맑은 고딕', size=13, bold=True)

ws3['A5'] = '"각 카테고리의 옷이 어떤 영화·인물·음악·시대와 연결되어 있는가" — 정보성 콘텐츠가 기술·역사 중심이라면, 문화 콘텐츠는 인물·서사·문화 중심.'
ws3['A5'].font = Font(name='맑은 고딕', size=10)
ws3['A5'].alignment = Alignment(wrap_text=True, vertical='top')
ws3.merge_cells('A5:C5')
ws3.row_dimensions[5].height = 40

# 카테고리별 예시
ws3['A7'] = '카테고리별 문화 앵글 예시'
ws3['A7'].font = Font(name='맑은 고딕', size=13, bold=True)

row = 8
ws3.cell(row=row, column=1, value='카테고리').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ws3.cell(row=row, column=2, value='문화 앵글').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ws3.cell(row=row, column=3, value='추천 콘텐츠').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
for col in range(1, 4):
    c = ws3.cell(row=row, column=col)
    c.fill = PatternFill('solid', fgColor='000000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border()
row += 1

culture_examples = [
    ('청자켓·데님자켓', '50s 블루칼라·영화 명배우', '영화 — 와일드 원 (말론 브란도, 1953), 이유 없는 반항 (제임스 딘, 1955), 스티브 맥퀸 (블리트 1968)'),
    ('바람막이+아노락', '아노락 흐름 — 원조부터 브릿팝까지', '이누이트 원조 → 1911 아문센 남극 정복 → 70s 등산 부활 → 90s 오아시스·갤러거 형제 브릿팝 / 영화 — 트레인스포팅 (1996, 90s 영국 청년)'),
    ('플리스', '벌목꾼·클라이밍 컬처', '노르웨이 벌목꾼 1961, 헬리한센 사연, 파타고니아 80s 캘리포니아 클라이머 / 영화 — 월터의 상상은 현실이 된다 (2013), 메루 (2015)'),
    ('패딩조끼', '90s 힙합·올드머니·캐릭터', '90s 뉴잉글랜드 프레피, 80s 등산 붐, 힙합 MV 패딩베스트 / 영화 — 더 로열 테넌바움 (웨스 앤더슨, 2001, 리치 테넌바움 트레이닝 룩)'),
    ('마운틴파카', '70s 백패킹 붐·자연 영화', '시에라 클럽, 미국 50개 주 종단 트레킹, 안셀 애덤스 사진 / 영화 — 인투 더 와일드 (2007, 알래스카), 와일드 (2014, PCT 트레킹)'),
    ('패딩', '80s 등산가·탐험가·등반 영화', '에디바우어 1936 Skyliner, 노르웨이 탐험대, 에베레스트 1953 / 영화 — 에베레스트 (2015, 1996 사고), 클리프행어, K2 (1991)'),
    ('모자', '데일리 아이콘·영화 캐릭터', '폴 뉴먼, 잭 케루악, 70s 트럭 드라이버 / 영화 — 포레스트 검프 (1994, 톰 행크스 트럭커캡), 캐스트어웨이 (2000), 머니볼 (브래드 피트)'),
    ('청바지 (데님)', '록밴드 무대·로드 무비', '리바이스 505 — 롤링 스톤즈, 라몬즈, 도어즈, 비치 보이즈 / 영화 — 이지 라이더 (1969), 텔마와 루이스, 영 원즈'),
    ('맨투맨·후드티', '록키·풋팰·90s 힙합', '영화 — 록키 (1976, 회색 트레이닝 후드), 굿 윌 헌팅 (1997) / 풋볼 컬리지 시대, 챔피언 90s 빅 사이즈'),
    ('필드·헌팅자켓', '밀리터리·영화·서브컬처', '영화 — 택시 드라이버 (1976, 트래비스 M-65), 풀 메탈 자켓 (1987), 디어 헌터 (1978) / 모드 vs 로커, 펑크 신'),
    ('코위찬·헤비스웨터', '원주민 문화·영화', '캐나다 코위찬 부족 손뜨개, 노르웨이 어부 / 영화 — 빅 르보스키 (1998, 더 듀드의 가디건), 더 미션 (1986)'),
    ('플란넬·체크셔츠', '90s 그런지·영화·미국 워크웨어', '영화 — 파이트 클럽 (1999, 브래드 피트), 캡틴 판타스틱 (2016), 트와일라잇 / 음악 — 너바나·펄잼·사운드가든 90s 시애틀 그런지 / 미국 농부·트럭 운전사·캠퍼 데일리'),
    ('코듀로이', '70s 영화·인디 캐릭터', '영화 — 더 로열 테넌바움, 라이프 어쿼틱 (웨스 앤더슨), 보이후드 (2014), 라스트 픽처 쇼 (1971) / 매튜 매커너히 70s 룩'),
]

for cat_name, angle, examples in culture_examples:
    ws3.cell(row=row, column=1, value=cat_name)
    ws3.cell(row=row, column=2, value=angle)
    ws3.cell(row=row, column=3, value=examples)
    for col in range(1, 4):
        c = ws3.cell(row=row, column=col)
        c.font = Font(name='맑은 고딕', size=10)
        c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws3.row_dimensions[row].height = 32
    row += 1

ws3.column_dimensions['A'].width = 22
ws3.column_dimensions['B'].width = 26
ws3.column_dimensions['C'].width = 60

# 레퍼런스 채널 (인스타) 섹션
row += 2
ws3.cell(row=row, column=1, value='참고 레퍼런스 채널 (문화·영화·서브컬처)')
ws3.cell(row=row, column=1).font = Font(name='맑은 고딕', size=13, bold=True)
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
row += 1

ws3.cell(row=row, column=1, value='문화 콘텐츠 기획·제작 시 참고할 인스타그램 레퍼런스 게시물. 톤·구성·시각 표현 벤치마킹용.')
ws3.cell(row=row, column=1).font = Font(name='맑은 고딕', size=10, color='525252')
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
ws3.cell(row=row, column=1).alignment = Alignment(wrap_text=True, vertical='top')
ws3.row_dimensions[row].height = 28
row += 2

ws3.cell(row=row, column=1, value='No.').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ws3.cell(row=row, column=2, value='유형').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ws3.cell(row=row, column=3, value='링크 (인스타그램)').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
for col in range(1, 4):
    c = ws3.cell(row=row, column=col)
    c.fill = PatternFill('solid', fgColor='000000')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = thin_border()
row += 1

ref_links = [
    ('R-01', '문화·영화·서브컬처', 'https://www.instagram.com/p/DYy9Su3j84p/'),
    ('R-02', '문화·영화·서브컬처', 'https://www.instagram.com/p/DYeQH9SD9gU/'),
    ('R-03', '문화·영화·서브컬처', 'https://www.instagram.com/p/DVgFynegdUI/'),
    ('R-04', '문화·영화·서브컬처', 'https://www.instagram.com/p/DY1eAXHGAxz/'),
    ('R-05', '문화·영화·서브컬처', 'https://www.instagram.com/p/DVIWDv9k5MS/'),
    ('R-06', '문화·영화·서브컬처', 'https://www.instagram.com/p/DSKFfIZASQP/'),
]

for no, kind, link in ref_links:
    ws3.cell(row=row, column=1, value=no)
    ws3.cell(row=row, column=2, value=kind)
    cell = ws3.cell(row=row, column=3, value=link)
    cell.hyperlink = link
    cell.font = Font(name='맑은 고딕', size=10, color='0563C1', underline='single')
    for col in range(1, 4):
        c = ws3.cell(row=row, column=col)
        if col != 3:
            c.font = Font(name='맑은 고딕', size=10)
        c.alignment = Alignment(horizontal='left' if col == 3 else 'center', vertical='center', wrap_text=True)
        c.border = thin_border()
    ws3.row_dimensions[row].height = 22
    row += 1

# 운영 메모
row += 1
ws3.cell(row=row, column=1, value='운영 메모')
ws3.cell(row=row, column=1).font = Font(name='맑은 고딕', size=12, bold=True)
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
row += 1

memos = [
    '• 정보성 콘텐츠가 "기술·역사·원단" 중심이라면, 문화 콘텐츠는 "인물·영화·음악·시대" 중심으로 차별',
    '• 영화 자료는 스틸컷 + 출연 연도 + 배우 이름을 함께 표기해 정보성 강화',
    '• 인용·이미지 사용 시 출처 표기 권장 (영화 캡처는 학술·리뷰 목적 인용 범위 내)',
    '• 발행 요일 — 매주 목요일 1회. 카테고리당 1회 (시즌 메인 발행 주차에 함께)',
    '• 콘텐츠 형식 — 카드뉴스 7~10장 권장. 짧은 영상(릴스)도 가능',
]
for memo in memos:
    ws3.cell(row=row, column=1, value=memo)
    ws3.cell(row=row, column=1).font = Font(name='맑은 고딕', size=10)
    ws3.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws3.row_dimensions[row].height = 22
    row += 1

# 저장
wb.save(OUT)
print(f'✓ xlsx 저장: {OUT}')
print(f'  Sheet 1: 카테고리 안내')
print(f'  Sheet 2: 발행 일정 ({len(CATEGORIES)}개 카테고리 × 6콘텐츠 = {len(CATEGORIES)*6}행)')
print(f'  Sheet 3: 문화 콘텐츠 가이드')
