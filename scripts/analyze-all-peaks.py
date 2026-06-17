"""
전 카테고리 연도별 정점 + 3년 평균 정점 일괄 분석
JSON 출력 → 슬라이드에 사용
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')
from openpyxl import load_workbook
from datetime import datetime, date
import statistics
import os
import json
import warnings
warnings.filterwarnings('ignore')

DOWNLOADS = r'C:\Users\wwkfl\Downloads'
OUT = r'C:\Users\wwkfl\Desktop\heavyduty-content-site\reports\peaks.json'

CATEGORIES = [
    ('padding',        '패딩 · 다운파카',     'datalab.xlsx',      '빈티지 패딩', '겨울', (10, 12)),
    ('mountain-parka', '마운틴파카',          'datalab (17).xlsx', '마운틴파카',  '겨울', (10, 12)),
    ('heavy-sweater',  '헤비스웨터 · 코위찬', 'datalab (2).xlsx',  '코위찬',     '겨울', (10, 12)),
    ('vest-jumper',    '점퍼 · 패딩베스트',   'datalab (3).xlsx',  '패딩베스트',  '가을', (10, 12)),
    ('corduroy',       '코듀로이',          'datalab (4).xlsx',  '코듀로이',    '겨울', (10, 12)),
    ('fleece',         '플리스 자켓',        'datalab (5).xlsx',  '플리스',     '가을', (10, 12)),
    ('sweatshirt',     '맨투맨 · 후드티',    'datalab (9).xlsx',  '맨투맨',     '가을',  (9, 11)),
    ('hunting-jacket', '헌팅 · 워크자켓',    'datalab (15).xlsx', '워크자켓',   '봄',     (2, 4)),
    ('shell-parka',    '쉘파카',            'datalab (1).xlsx',  '쉘파카',     '봄',     (2, 4)),
    ('denim',          '데님 · 청바지',      'datalab (10).xlsx', '청바지',     '봄',     (2, 4)),
    ('anorak-coach',   '아노락 · 코치자켓',   'datalab (7).xlsx',  '아노락',     '봄',     (2, 4)),
    ('work-shirt',     '워크셔츠',          'datalab (14).xlsx', '폴로셔츠',   '봄',     (3, 5)),
    ('chino-cargo',    '치노 · 카고 · 워크팬츠','datalab (11).xlsx','치노팬츠',   '봄',     (2, 4)),
    ('tshirt',         '반팔 티셔츠',        'datalab (12).xlsx', '반팔티',     '여름',   (4, 6)),
    ('shorts',         '반바지',            'datalab (13).xlsx', '반바지',     '여름',   (5, 7)),
    ('denim-jacket',   '청자켓',            'datalab (16).xlsx', '청자켓',     '봄 메인 + 가을 서브', (9, 11)),
]

def load_keyword_series(fname, target):
    wb = load_workbook(os.path.join(DOWNLOADS, fname), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r and r[0] == '날짜')
    header = rows[header_idx]
    n_pairs = len(header) // 2
    for p in range(n_pairs):
        kw = header[p*2+1]
        if kw != target: continue
        s = {}
        for r in rows[header_idx+1:]:
            if r[p*2] and r[p*2+1] is not None:
                d = r[p*2].date() if hasattr(r[p*2],'date') else datetime.strptime(str(r[p*2])[:10],'%Y-%m-%d').date()
                s[d] = float(r[p*2+1])
        return s
    return None

def ma7(daily):
    if not daily: return {}
    dates = sorted(daily.keys())
    out = {}
    for i, dt in enumerate(dates):
        vals = [daily[dates[j]] for j in range(max(0,i-6), i+1)]
        out[dt] = statistics.mean(vals)
    return out

def yearly_peak_in_window(ma, year, m_start, m_end):
    window = sorted([(d, v) for d, v in ma.items() if d.year == year and m_start <= d.month <= m_end])
    if not window: return None
    peak_d, peak_v = max(window, key=lambda x: x[1])
    thresh = peak_v * 0.9
    peak_idx = next(i for i, (d, _) in enumerate(window) if d == peak_d)
    si = peak_idx
    while si > 0 and window[si-1][1] >= thresh:
        si -= 1
    return {'peak_date': peak_d.isoformat(), 'peak_md': peak_d.strftime('%m/%d'), 'peak_val': round(peak_v,1), 'start_date': window[si][0].isoformat(), 'start_md': window[si][0].strftime('%m/%d')}

def avg_md_peak(ma, m_start, m_end):
    """3년 평균 (월·일 → 평균)"""
    by_md = {}
    for d, v in ma.items():
        if d.year not in [2023,2024,2025]: continue
        if not (m_start <= d.month <= m_end): continue
        by_md.setdefault((d.month, d.day), []).append(v)
    avg = {k: statistics.mean(vs) for k, vs in by_md.items()}
    items = sorted(avg.items())
    peak_md, peak_v = max(items, key=lambda x: x[1])
    thresh = peak_v * 0.9
    pi = next(i for i, (k,_) in enumerate(items) if k == peak_md)
    si = pi
    while si > 0 and items[si-1][1] >= thresh:
        si -= 1
    return {
        'peak_md': f'{peak_md[0]}/{peak_md[1]}',
        'peak_val': round(peak_v, 1),
        'start_md': f'{items[si][0][0]}/{items[si][0][1]}'
    }

def week_label_of(md_str):
    """'10/7' → '10월 1주차'"""
    m, d = map(int, md_str.split('/'))
    w = 1 if d <= 7 else 2 if d <= 14 else 3 if d <= 21 else 4
    return f'{m}월 {w}주차'

result = {}
for cat_id, name, fname, kw, season, (ms, me) in CATEGORIES:
    daily = load_keyword_series(fname, kw)
    if not daily:
        result[cat_id] = {'error': f'keyword {kw} not found'}
        continue
    ma = ma7(daily)
    by_year = {}
    for year in [2023, 2024, 2025]:
        p = yearly_peak_in_window(ma, year, ms, me)
        if p: by_year[year] = p
    avg_p = avg_md_peak(ma, ms, me)
    # 분산
    from datetime import date as dt
    peak_doys = [dt.fromisoformat(p['peak_date']).timetuple().tm_yday for p in by_year.values()]
    spread = (max(peak_doys) - min(peak_doys)) if len(peak_doys) >= 2 else 0
    consistency = '일관' if spread <= 14 else '보통' if spread <= 30 else '들쭉'
    result[cat_id] = {
        'name': name,
        'keyword': kw,
        'season': season,
        'window': f'{ms}월~{me}월',
        'by_year': by_year,
        'avg_3yr': avg_p,
        'avg_peak_week': week_label_of(avg_p['peak_md']),
        'spread_days': spread,
        'consistency': consistency
    }

with open(OUT, 'w', encoding='utf-8') as f:
    json.dump(result, f, ensure_ascii=False, indent=2)

# 요약 출력
print(f'\n전 카테고리 정점 분석 — {len(result)}개')
print(f'{"카테고리":<22} {"키워드":<12} {"23 정점":<8} {"24 정점":<8} {"25 정점":<8} {"3년평균":<12} {"일관성":<6}')
print('─' * 95)
for cat_id, r in result.items():
    if 'error' in r:
        print(f'{r.get("name","?"):<22}  ✗  {r["error"]}')
        continue
    y23 = r['by_year'].get(2023, {}).get('peak_md', '-')
    y24 = r['by_year'].get(2024, {}).get('peak_md', '-')
    y25 = r['by_year'].get(2025, {}).get('peak_md', '-')
    print(f'{r["name"]:<22} {r["keyword"]:<12} {y23:<8} {y24:<8} {y25:<8} {r["avg_3yr"]["peak_md"]:<5} ({r["avg_peak_week"]:<8}) {r["consistency"]:<6}')

print(f'\n저장: {OUT}')
