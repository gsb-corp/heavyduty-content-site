"""
헤비듀티 카테고리 3년 트렌드 곡선 그래프 생성
- 각 카테고리별 1장 PNG
- 2023/2024/2025 3개 곡선 + 연도별 정점 표시
- 미니멀 흑백 디자인 (사이트와 통일)
"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from datetime import datetime, date
import statistics
import os
import warnings
warnings.filterwarnings('ignore')

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from matplotlib import font_manager, rc

# 한글 폰트 (윈도우 Malgun Gothic)
plt.rcParams['font.family'] = 'Malgun Gothic'
plt.rcParams['axes.unicode_minus'] = False

DOWNLOADS = r'C:\Users\wwkfl\Downloads'
OUT_DIR = r'C:\Users\wwkfl\Desktop\heavyduty-content-site\reports\graphs'

# 카테고리 매핑 (v13 데이터 + 청자켓 추가)
# (id, 한국명, 데이터랩 파일, 대표 키워드, 시즌)
CATEGORIES = [
    ('padding',       '패딩 · 다운파카',     'datalab.xlsx',      '빈티지 패딩', '겨울'),
    ('mountain-parka', '마운틴파카 · 밀리터리', 'datalab (17).xlsx', '마운틴파카',  '겨울'),
    ('heavy-sweater', '헤비스웨터 · 코위찬',  'datalab (2).xlsx',  '코위찬',     '겨울'),
    ('vest-jumper',   '점퍼 · 패딩베스트',   'datalab (3).xlsx',  '패딩베스트',  '가을'),
    ('corduroy',      '코듀로이',          'datalab (4).xlsx',  '코듀로이',    '겨울'),
    ('fleece',        '플리스 자켓',        'datalab (5).xlsx',  '플리스',     '가을'),
    ('sweatshirt',    '맨투맨 · 후드티',    'datalab (9).xlsx',  '맨투맨',     '가을'),
    ('hunting-jacket','헌팅 · 워크 · 필드자켓','datalab (15).xlsx', '워크자켓',   '봄'),
    ('shell-parka',   '쉘파카',            'datalab (1).xlsx',  '쉘파카',     '봄'),
    ('denim',         '데님 · 청바지',      'datalab (10).xlsx', '청바지',     '봄'),
    ('anorak-coach',  '아노락 · 코치자켓',   'datalab (7).xlsx',  '아노락',     '봄'),
    ('work-shirt',    '워크셔츠 · 체크셔츠', 'datalab (14).xlsx', '폴로셔츠',   '봄'),
    ('chino-cargo',   '치노 · 카고 · 워크팬츠','datalab (11).xlsx','치노팬츠',   '봄'),
    ('tshirt',        '반팔 티셔츠',        'datalab (12).xlsx', '반팔티',     '여름'),
    ('shorts',        '반바지 · 카고쇼츠',   'datalab (13).xlsx', '반바지',     '여름'),
    ('denim-jacket',  '청자켓 · 데님자켓',   'datalab (16).xlsx', '청자켓',     '봄(가을 서브)'),
]

def load_series(fname, target_keyword):
    """다중 키워드 파일에서 특정 키워드 시계열 추출"""
    wb = load_workbook(os.path.join(DOWNLOADS, fname), data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header_idx = next(i for i, r in enumerate(rows) if r and r[0] == '날짜')
    header = rows[header_idx]
    # 키워드별 컬럼 짝 찾기 (날짜 N, 값 N+1)
    series = {}
    n_pairs = len(header) // 2
    for p in range(n_pairs):
        kw = header[p*2+1]
        if not kw: continue
        col_date = p*2
        col_val = p*2+1
        s = {}
        for r in rows[header_idx+1:]:
            if r[col_date] and r[col_val] is not None:
                d = r[col_date].date() if hasattr(r[col_date],'date') else datetime.strptime(str(r[col_date])[:10],'%Y-%m-%d').date()
                s[d] = float(r[col_val])
        series[kw] = s
    return series.get(target_keyword)

def ma7(daily):
    """7일 이동평균"""
    if not daily: return {}
    dates = sorted(daily.keys())
    out = {}
    for i, dt in enumerate(dates):
        vals = [daily[dates[j]] for j in range(max(0,i-6), i+1)]
        out[dt] = statistics.mean(vals)
    return out

def yearly_peak(ma, year):
    """특정 연도의 정점 일자, 값, 90% 진입일"""
    year_data = sorted([(d, v) for d, v in ma.items() if d.year == year])
    if not year_data: return None
    peak_d, peak_v = max(year_data, key=lambda x: x[1])
    return {'peak_date': peak_d, 'peak_val': peak_v}

def draw_graph(cat_id, cat_name, fname, target_kw, season):
    print(f'  ▸ {cat_name} ({target_kw})')
    daily = load_series(fname, target_kw)
    if not daily:
        print(f'    ✗ {target_kw} 시계열 없음')
        return None
    ma = ma7(daily)

    # 연도별 정점
    peaks = {y: yearly_peak(ma, y) for y in [2023, 2024, 2025]}

    # 그래프
    fig, ax = plt.subplots(figsize=(13, 5.2), dpi=180)
    fig.patch.set_facecolor('white')
    ax.set_facecolor('white')

    colors = {2023: '#BDBDBD', 2024: '#6B6B6B', 2025: '#000000'}
    linewidths = {2023: 2.0, 2024: 2.6, 2025: 3.6}

    # 각 연도 곡선 (월/일을 X축으로 통합)
    for year in [2023, 2024, 2025]:
        year_data = sorted([(d, v) for d, v in ma.items() if d.year == year])
        if not year_data: continue
        # X축을 1/1부터의 일수로 통일
        xs = [(d - date(year, 1, 1)).days for d, v in year_data]
        ys = [v for d, v in year_data]
        ax.plot(xs, ys, color=colors[year], linewidth=linewidths[year], label=f'{year}', alpha=0.95)

        # 정점 표시
        if peaks[year]:
            p_x = (peaks[year]['peak_date'] - date(year, 1, 1)).days
            p_y = peaks[year]['peak_val']
            ax.scatter([p_x], [p_y], color=colors[year], s=110, zorder=5, edgecolors='white', linewidth=2)
            # 라벨
            md = peaks[year]['peak_date'].strftime('%m/%d')
            ax.annotate(f'{md}',
                xy=(p_x, p_y),
                xytext=(p_x, p_y + max(ys)*0.08),
                ha='center', fontsize=11.5, fontweight='bold',
                color=colors[year])

    # X축: 월
    month_starts = [(date(2024, m, 1) - date(2024, 1, 1)).days for m in range(1, 13)]
    month_labels = ['1월', '2월', '3월', '4월', '5월', '6월', '7월', '8월', '9월', '10월', '11월', '12월']
    ax.set_xticks(month_starts)
    ax.set_xticklabels(month_labels, fontsize=12, color='#000', fontweight='bold')
    ax.set_xlim(0, 366)

    # Y축
    ax.tick_params(axis='y', labelsize=11, colors='#525252')
    ax.set_ylabel('검색량', fontsize=11, color='#525252', fontweight='bold')

    # 보조선
    ax.grid(True, axis='y', linestyle=':', linewidth=0.7, color='#BBBBBB', alpha=0.8)
    ax.set_axisbelow(True)

    # 테두리
    for spine_name, spine in ax.spines.items():
        if spine_name in ['top', 'right']:
            spine.set_visible(False)
        else:
            spine.set_color('#000')
            spine.set_linewidth(1.5)

    # 제목 + 부제 (좌측 상단)
    ax.text(0, 1.18, cat_name,
        transform=ax.transAxes, fontsize=20, fontweight='bold', color='#000', ha='left', va='top')
    ax.text(0, 1.06, f'키워드: {target_kw}   ·   메인 시즌: {season}',
        transform=ax.transAxes, fontsize=12, color='#525252', ha='left', va='top')

    # 범례
    legend = ax.legend(loc='upper right', frameon=False, fontsize=13, ncol=3, columnspacing=2,
        bbox_to_anchor=(1, 1.13))
    for t in legend.get_texts():
        t.set_color('#000')
        t.set_fontweight('bold')

    plt.tight_layout()
    out_path = os.path.join(OUT_DIR, f'{cat_id}.png')
    plt.savefig(out_path, dpi=180, bbox_inches='tight', facecolor='white')
    plt.close()
    return out_path

print('━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━')
print('카테고리 곡선 그래프 생성 (전 카테고리)')
print('━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━')

os.makedirs(OUT_DIR, exist_ok=True)

generated = []
for cat_id, cat_name, fname, target_kw, season in CATEGORIES:
    path = draw_graph(cat_id, cat_name, fname, target_kw, season)
    if path: generated.append((cat_id, cat_name, path))

print(f'\n✓ 완료: {len(generated)}장')
print(f'  저장 위치: {OUT_DIR}')
