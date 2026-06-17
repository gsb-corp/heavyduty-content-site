"""운영 흐름 예시 시트만 갱신 — P-11 발주 + 단계 재정의 반영"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

XLSX = r'C:\Users\wwkfl\Desktop\heavyduty-content-site\reports\content-calendar\콘텐츠_캘린더.xlsx'

def thin_border():
    return Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

wb = load_workbook(XLSX)
print(f'시트 목록: {wb.sheetnames}')

if '운영 흐름 예시' in wb.sheetnames:
    del wb['운영 흐름 예시']
    print('  ✓ 기존 "운영 흐름 예시" 시트 삭제')

# 운영 흐름 예시 시트 — Sheet 3 (단계별 데드라인) 다음 위치
ws = wb.create_sheet('운영 흐름 예시', 3)

ws['A1'] = '운영 흐름 예시 — 청자켓·데님자켓 (정점 10/7 = 10월 1주차)'
ws['A1'].font = Font(name='맑은 고딕', size=16, bold=True)
ws.merge_cells('A1:M1')

ws['A2'] = '발주 2주(P-11~P-10) → 배송 4주(P-9~P-6) → 케어·업로드 3주(P-7~P-5) → 기획·제작 2주(P-5~P-4) → 발행 2주(P-3~P-2) → 정점.'
ws['A2'].font = Font(name='맑은 고딕', size=10, color='525252')
ws.merge_cells('A2:M2')

# 주차 헤더 (P-11 ~ P, 11개) — 리마인드 제거로 P+ 영역 삭제
WEEKS = [
    ('7월\n3주차', 'P-11주'),    # 0
    ('7월\n4주차', 'P-10주'),    # 1
    ('8월\n1주차', 'P-9주'),     # 2
    ('8월\n2주차', 'P-8주'),     # 3
    ('8월\n3주차', 'P-7주'),     # 4
    ('8월\n4주차', 'P-6주'),     # 5
    ('9월\n1주차', 'P-5주'),     # 6
    ('9월\n2주차', 'P-4주'),     # 7
    ('9월\n3주차', 'P-3주'),     # 8
    ('9월\n4주차', 'P-2주'),     # 9
    ('10월\n1주차','🎯 정점'),    # 10
]

# 단계별 활동 주차 (사용자 확정 — 모든 카테고리 공통 흐름)
# (단계명, [활동 인덱스], 색상, 설명)
STAGES = [
    ('① 사입 발주',           [0, 1],       'D32F2F', '🔴 가장 중요한 데드라인 — 해외 발주 (P-11~P-10주, 2주 분산)'),
    ('② 배송 기간',            [2, 3, 4, 5], '0070C0', '발주 후 배송·통관·도착 (P-9~P-6주, 4주). 안전 여유 포함.'),
    ('③ 제품 케어 및 업로드', [4, 5, 6],    '00B050', '세탁·검수·제품 촬영·매장 등록 (P-7~P-5주, 3주)'),
    ('④ 콘텐츠 기획·제작',    [6, 7],       'FFC000', '서사·카드뉴스·릴스·문화 콘텐츠 편집 (P-5~P-4주, 2주)'),
    ('⑤ 발행 (6콘텐츠)',      [8, 9],       'FF6B6B', '🟡 월·화·수·목(문화)·금·일 6콘텐츠 분산 발행 (P-3~P-2주, 2주)'),
    ('🎯 정점 (P)',            [10],         '000000', '검색량 90% 진입 첫날 — 골든타임 시작 (정점 ±2주 4주간)'),
]

# 헤더 작성
header_row = 4
ws.cell(row=header_row, column=1, value='단계').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ws.cell(row=header_row, column=1).fill = PatternFill('solid', fgColor='000000')
ws.cell(row=header_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
ws.cell(row=header_row, column=1).border = thin_border()

for i, (week_label, p_label) in enumerate(WEEKS):
    col = i + 2
    c = ws.cell(row=header_row, column=col, value=f'{week_label}\n({p_label})')
    c.font = Font(name='맑은 고딕', size=9, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='000000' if '정점' in p_label else '404040')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_border()

desc_col = len(WEEKS) + 2
ws.cell(row=header_row, column=desc_col, value='설명').font = Font(name='맑은 고딕', size=11, bold=True, color='FFFFFF')
ws.cell(row=header_row, column=desc_col).fill = PatternFill('solid', fgColor='000000')
ws.cell(row=header_row, column=desc_col).alignment = Alignment(horizontal='center', vertical='center')
ws.cell(row=header_row, column=desc_col).border = thin_border()

ws.row_dimensions[header_row].height = 45

# 단계별 행
for stage_idx, (stage_name, active_weeks, color, desc) in enumerate(STAGES):
    row = header_row + 1 + stage_idx
    # 단계 이름
    c = ws.cell(row=row, column=1, value=stage_name)
    c.font = Font(name='맑은 고딕', size=10, bold=True)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = thin_border()

    # 주차별 셀
    for i in range(len(WEEKS)):
        col = i + 2
        c = ws.cell(row=row, column=col)
        if i in active_weeks:
            c.value = '■'
            c.font = Font(name='맑은 고딕', size=14, bold=True, color='FFFFFF')
            c.fill = PatternFill('solid', fgColor=color)
            c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()

    # 설명
    c = ws.cell(row=row, column=desc_col, value=desc)
    c.font = Font(name='맑은 고딕', size=10)
    c.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True, indent=1)
    c.border = thin_border()

    ws.row_dimensions[row].height = 28

# 컬럼 너비
ws.column_dimensions['A'].width = 22
for i in range(len(WEEKS)):
    ws.column_dimensions[get_column_letter(i+2)].width = 9
ws.column_dimensions[get_column_letter(desc_col)].width = 60

# 채택 안내 박스
row = header_row + len(STAGES) + 3
ws.cell(row=row, column=1, value='✅ 채택된 발행 기준 — A안 (정점 -2주)')
ws.cell(row=row, column=1).font = Font(name='맑은 고딕', size=13, bold=True, color='FFFFFF')
ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='000000')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=desc_col)
ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[row].height = 26
row += 1

guidance = [
    '• 사입 발주 (P-11~P-10주, 2주) — 카테고리별 분산 가능. P-10주까지 발주 마감.',
    '• 배송 기간 (P-9~P-6주, 4주) — 해외 배송 + 통관 + 안전 여유 포함. P-6주에 사무실 도착.',
    '• 제품 케어·업로드 (P-7~P-5주, 3주) — 도착 직후부터 시작. 세탁·촬영·매장 등록.',
    '• 콘텐츠 기획·제작 (P-5~P-4주, 2주) — 제품 사진 활용 카드뉴스·릴스·문화 콘텐츠 편집.',
    '• 발행 (P-3~P-2주, 2주) — 정점 직전 2주에 6콘텐츠 분산 발행. 정점에 가까이 노출.',
    '• 모든 카테고리 동일 흐름 적용 — 정점 시점만 다름 (각 카테고리 정점 기준 단계 자동 계산).',
]
for g in guidance:
    ws.cell(row=row, column=1, value=g)
    ws.cell(row=row, column=1).font = Font(name='맑은 고딕', size=10.5)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=desc_col)
    ws.row_dimensions[row].height = 22
    row += 1

# 저장
wb.save(XLSX)
print(f'\n✓ "운영 흐름 예시" 시트 갱신 완료')
print(f'  P-11 사입 발주 + P-9 자동 도착 반영')
print(f'  발행 기준 A안 (정점 -2주) 채택 명시')
print(f'  다른 시트(사용자 수정 포함)는 그대로 보존됨')
