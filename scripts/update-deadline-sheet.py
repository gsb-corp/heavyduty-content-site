"""단계별 데드라인 시트만 갱신 — 다른 시트(사용자 수정) 보존"""
import sys
sys.stdout.reconfigure(encoding='utf-8')

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date, timedelta

XLSX = r'C:\Users\wwkfl\Desktop\heavyduty-content-site\reports\content-calendar\콘텐츠_캘린더.xlsx'

# 카테고리 (gen-content-calendar.py와 동일)
CATEGORIES = [
    (1,  '플란넬·체크셔츠',      'C 서브',     '9/3',    date(2026, 9, 3)),
    (2,  '마운틴파카 (60/40)',   'A 시즌메인', '9/28',   date(2026, 9, 28)),
    (3,  '바람막이+아노락',      'A 시즌메인', '10/1',   date(2026, 10, 1)),
    (4,  '청자켓·데님자켓',      'A 시즌메인', '10/7',   date(2026, 10, 7)),
    (5,  '필드·헌팅자켓',        'C 메인추가', '환절기',  None),
    (6,  '맨투맨·후드티',        'C 메인추가', '10/26',  date(2026, 10, 26)),
    (7,  '플리스',               'A 시즌메인', '11/9',   date(2026, 11, 9)),
    (8,  '패딩조끼',             'A 시즌메인', '11/16',  date(2026, 11, 16)),
    (9,  '코위찬·헤비스웨터',    'C 서브',     '11/28',  date(2026, 11, 28)),
    (10, '패딩',                 'A 시즌메인', '12/1',   date(2026, 12, 1)),
    (11, '코듀로이',             'C 서브',     '12/1',   date(2026, 12, 1)),
]

def date_to_week_with_range(d):
    if d is None:
        return '연중 (자유)'
    weekday = d.weekday()
    monday = d - timedelta(days=weekday)
    sunday = monday + timedelta(days=6)
    day = d.day
    if day <= 7: w = 1
    elif day <= 14: w = 2
    elif day <= 21: w = 3
    else: w = 4
    return f'{d.month}월 {w}주차\n({monday.month}/{monday.day}~{sunday.month}/{sunday.day})'

def stage_dates(peak):
    """사용자 확정 운영 흐름 — 각 단계 마감일 (그 시점까지 끝나야 함)."""
    if peak is None: return None
    return {
        'order':    peak - timedelta(weeks=10),  # ① 사입 발주 마감 (발주 P-11~P-10, 2주)
        'shipping': peak - timedelta(weeks=6),   # ② 배송 완료 (P-9~P-6, 4주)
        'care':     peak - timedelta(weeks=5),   # ③ 케어·업로드 완료 (P-7~P-5, 3주)
        'creative': peak - timedelta(weeks=4),   # ④ 기획·제작 완료 (P-5~P-4, 2주)
        'launch':   peak - timedelta(weeks=3),   # ⑤ 발행 시작 (P-3~P-2, 2주)
    }

def thin_border():
    return Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

# 1. 기존 파일 로드
wb = load_workbook(XLSX)
print(f'시트 목록: {wb.sheetnames}')

# 2. 단계별 데드라인 시트 삭제 (있으면)
if '단계별 데드라인' in wb.sheetnames:
    del wb['단계별 데드라인']
    print('  ✓ 기존 "단계별 데드라인" 시트 삭제')

# 3. 새로 추가 — Sheet 2 (발행 일정) 다음 위치에
ws = wb.create_sheet('단계별 데드라인', 2)

ws['A1'] = '단계별 데드라인 — 11개 시즌 카테고리 정점 -2주 발행 기준'
ws['A1'].font = Font(name='맑은 고딕', size=16, bold=True)
ws.merge_cells('A1:H1')

ws['A2'] = '각 셀 = 해당 단계의 마감 데드라인. 발주 2주 → 배송 4주 → 케어 3주 → 제작 2주 → 발행 2주 → 정점.'
ws['A2'].font = Font(name='맑은 고딕', size=10, color='525252')
ws.merge_cells('A2:H2')

# 헤더 (8컬럼)
row = 4
headers = ['#', '카테고리', '정점',
           '① 사입 발주 마감\n(P-10주)\n발주 기간 P-11~P-10 (2주)',
           '② 배송 완료\n(P-6주)\n배송 P-9~P-6 (4주)',
           '③ 케어·업로드 완료\n(P-5주)\n케어 P-7~P-5 (3주)',
           '④ 기획·제작 완료\n(P-4주)\n제작 P-5~P-4 (2주)',
           '⑤ 발행 시작\n(P-3주)\n발행 P-3~P-2 (2주)']
for col, h in enumerate(headers, 1):
    c = ws.cell(row=row, column=col, value=h)
    c.font = Font(name='맑은 고딕', size=9, bold=True, color='FFFFFF')
    c.fill = PatternFill('solid', fgColor='000000')
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = thin_border()
ws.row_dimensions[row].height = 60
row += 1

# 데이터 행
for cat in CATEGORIES:
    seq, name, role, peak, peak_date = cat
    stages = stage_dates(peak_date)

    # 색상 — A·C메인 = 진한 주황 / C 서브 = 연회색
    is_confirmed = 'A' in role or 'C 메인' in role
    fill_color = 'FFB84D' if is_confirmed else 'F5F5F5'

    ws.cell(row=row, column=1, value=seq)
    ws.cell(row=row, column=2, value=name)
    ws.cell(row=row, column=3, value=peak)

    if stages:
        ws.cell(row=row, column=4, value=date_to_week_with_range(stages['order']))
        ws.cell(row=row, column=5, value=date_to_week_with_range(stages['shipping']))
        ws.cell(row=row, column=6, value=date_to_week_with_range(stages['care']))
        ws.cell(row=row, column=7, value=date_to_week_with_range(stages['creative']))
        ws.cell(row=row, column=8, value=date_to_week_with_range(stages['launch']))

        for col in range(1, 9):
            c = ws.cell(row=row, column=col)
            c.font = Font(name='맑은 고딕', size=9, bold=is_confirmed or (col == 8))
            # ① 발주 마감 = 빨강 강조, ⑤ 발행 = 노랑 강조
            if col == 4:
                c.fill = PatternFill('solid', fgColor='FFD6D6')  # 사입 발주 마감
            elif col == 8:
                c.fill = PatternFill('solid', fgColor='FFFFCC')  # 발행
            else:
                c.fill = PatternFill('solid', fgColor=fill_color)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin_border()
    else:
        # 필드자켓 등 정점 없는 카테고리 — 셀 병합 + 안내
        for col in range(1, 4):
            c = ws.cell(row=row, column=col)
            c.font = Font(name='맑은 고딕', size=9, bold=is_confirmed)
            c.fill = PatternFill('solid', fgColor=fill_color)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border = thin_border()
        ws.cell(row=row, column=4, value='청자켓과 비슷한 시기, 컨셉트 공유하는 제품군 — 10월 3주차 빈 구간 추천. 정점 데이터 없어 자유 배치.')
        ws.merge_cells(start_row=row, start_column=4, end_row=row, end_column=8)
        c = ws.cell(row=row, column=4)
        c.font = Font(name='맑은 고딕', size=10, italic=True, color='525252')
        c.fill = PatternFill('solid', fgColor='FFF4E0')
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = thin_border()

    ws.row_dimensions[row].height = 50
    row += 1

# 운영 원칙 박스
row += 2
ws.cell(row=row, column=1, value='💡 데드라인 운영 원칙')
ws.cell(row=row, column=1).font = Font(name='맑은 고딕', size=13, bold=True, color='FFFFFF')
ws.cell(row=row, column=1).fill = PatternFill('solid', fgColor='000000')
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1)
ws.row_dimensions[row].height = 26
row += 1

principles = [
    '• ① 사입 발주 마감 (P-10주, 빨강) = 가장 중요한 데드라인. 발주는 P-11~P-10주 2주 동안 분산, P-10주까지 끝내야 함.',
    '• ② 배송 기간 (P-9~P-6, 4주) — 해외 배송 + 통관 + 안전 여유 포함. P-6주까지 우리 사무실 도착.',
    '• ③ 케어·업로드 (P-7~P-5, 3주) — 도착 직후부터 검수·세탁·제품 촬영·매장 등록.',
    '• ④ 기획·제작 (P-5~P-4, 2주) — 제품 사진 활용해 6콘텐츠(카드뉴스·릴스·문화) 편집.',
    '• ⑤ 발행 (P-3~P-2, 2주, 노랑) — 정점 직전 2주 동안 6콘텐츠 분산 발행. 정점에 가장 가까이 노출.',
    '• 모자(포시즌)·청바지(포시즌)는 시즌 시퀀스와 무관하게 사입 입고 시점에 맞춰 자유롭게 운영.',
    '• 필드·헌팅자켓은 정점 데이터(환절기)가 명확하지 않아 자유 배치. 청자켓 시즌과 결 공유 → 10월 3주차 빈 구간 추천.',
]
for p in principles:
    ws.cell(row=row, column=1, value=p)
    ws.cell(row=row, column=1).font = Font(name='맑은 고딕', size=10.5)
    ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    ws.row_dimensions[row].height = 22
    row += 1

# 컬럼 너비 (8컬럼)
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 12
for c_letter in ['D', 'E', 'F', 'G', 'H']:
    ws.column_dimensions[c_letter].width = 22

# 저장
wb.save(XLSX)
print(f'\n✓ "단계별 데드라인" 시트 갱신 완료')
print(f'  다른 시트(사용자 수정 포함)는 그대로 보존됨')
